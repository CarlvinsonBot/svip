import os
import telegram
import openpyxl
from io import BytesIO
import requests
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters
from telegram import InlineKeyboardButton, InlineKeyboardMarkup

# 机器人 token
TOKEN = '5983348762:AAGyhzIL29C0lal3zjgAsH9JeWutoor_W-A'

# GitHub 文件地址
file_url = 'https://raw.githubusercontent.com/CarlvinsonBot/HelloWorld/blob/fb8d1c75f18373af435bc6db45ee87362ec7e8e0/PTS%E8%AE%A1%E7%AE%97%E8%A1%A8%E6%A0%BC.xlsx'

# 创建一个机器人实例
bot = telegram.Bot(token=TOKEN)

# 下载文件并读取数据
def read_data(update, context):
    response = requests.get(file_url)
    file_stream = BytesIO(response.content)
    workbook = openpyxl.load_workbook(file_stream)
    sheet = workbook.active
    for row in sheet.iter_rows():
        for cell in row:
            context.bot.send_message(chat_id=update.effective_chat.id, text=cell.value)

# 创建一个更新器实例
updater = Updater(TOKEN, use_context=True)
dispatcher = updater.dispatcher
# 定义 /start 命令的处理程序
def start(update, context):
    context.bot.send_message(chat_id=update.effective_chat.id, text="欢迎使用本机器人！请按照格式发送数据。")
# 将命令处理程序添加到调度程序中
start_handler = CommandHandler('start', start)
dispatcher.add_handler(start_handler)

# 加载工作簿
url = 'https://raw.githubusercontent.com/CarlvinsonBot/HelloWorld/PTS计算表格.xlsx'
response = requests.get(url)
workbook = load_workbook(filename=response.content)

# 获得工作表1和工作表2
worksheet1 = workbook['自定义四层账户额度']
worksheet2 = workbook['总投入平均分配额度']

# 定义接收数据的信息处理程序
def receive_data(update, context):
    # 从信息中获取数据
    data = update.message.text.split()

    # 检查数据的格式和它的有效性
    if len(data) == 6 and all(int(x) >= 200 for x in data[:4]) and all(int(x) >= 1 for x in data[4:]):
        # 如果数据格式为5组数，更新工作表1
        worksheet1['B2'] = data[0]
        worksheet1['B3'] = data[1]
        worksheet1['B4'] = data[2]
        worksheet1['B5'] = data[3]
        worksheet1['D5'] = data[4]
        worksheet1['D1'] = data[5]

        # 保存工作簿
        workbook.save('PTS计算表格.xlsx')

        # 发送消息附带按钮
        buttons = [
            [InlineKeyboardButton("目录1", callback_data="menu1"), InlineKeyboardButton("目录2", callback_data="menu2")],
            [InlineKeyboardButton("返回", callback_data="back")]
        ]
        reply_markup = InlineKeyboardMarkup(buttons)
        context.bot.send_message(chat_id=update.effective_chat.id, text=f"C24: {data[4]}\nE24: {data[5]}", reply_markup=reply_markup)
    elif len(data) == 3 and int(data[0]) >= 800 and all(int(x) >= 1 for x in data[1:]):
        # 如果数据格式为2组数，更新工作表2
        worksheet2['B1'] = data[0]
        worksheet2['D4'] = data[1]
        worksheet2['D1'] = data[2]

        # 保存工作簿
        workbook.save('PTS计算表格.xlsx')

        # 发送消息附带按钮
        buttons = [
            [InlineKeyboardButton("目录1", callback_data="menu1"), InlineKeyboardButton("目录2", callback_data="menu2")],
            [InlineKeyboardButton("返回", callback_data="back")]
        ]
        reply_markup = InlineKeyboardMarkup(buttons)
        context.bot.send_message(chat_id=update.effective_chat.id, text=f"C24: {data[1]}\nE24: {data[2]}", reply_markup=reply_markup)
    else:
        context.bot.send_message(chat_id=update.effective_chat.id, text="数据格式不正确或数据无效！请按照正确格式发送数据。")
   
# 处理回复消息中的按钮
def handle_callback(update, context):
    query = update.callback_query
    chat_id = query.message.chat_id
    message_id = query.message.message_id
    if query.data == 'button1':
        keyboard = [[InlineKeyboardButton("目录1", callback_data='directory1'),
                     InlineKeyboardButton("目录2", callback_data='directory2'),
                     InlineKeyboardButton("目录3", callback_data='directory3')],
                    [InlineKeyboardButton("返回", callback_data='back')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.send_message(chat_id=chat_id, message_id=message_id,
                                      text="以下是目录", reply_markup=reply_markup)
    elif query.data == 'directory1':
        keyboard = [[InlineKeyboardButton("返回", callback_data='back')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.send_message(chat_id=chat_id, message_id=message_id,
                                      text="这是目录1的内容", reply_markup=reply_markup)
    elif query.data == 'directory2':
        keyboard = [[InlineKeyboardButton("返回", callback_data='back')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.send_message(chat_id=chat_id, message_id=message_id,
                                      text="这是目录2的内容", reply_markup=reply_markup)
    elif query.data == 'directory3':
        keyboard = [[InlineKeyboardButton("返回", callback_data='back')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.send_message(chat_id=chat_id, message_id=message_id,
                                      text="这是目录3的内容", reply_markup=reply_markup)
    elif query.data == 'back':
        keyboard = [[InlineKeyboardButton("按钮1", callback_data='button1'),
                     InlineKeyboardButton("按钮1", url='https://t.me/VinsonChannel')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        c24_value = worksheet.acell('C24').value
        e24_value = worksheet.acell('E24').value
        message_text = f"C24: {c24_value}\nE24: {e24_value}"
        context.bot.send_message(chat_id=chat_id, message_id=message_id,
                                      text=message_text, reply_markup=reply_markup)      
# 启动机器人
updater.start_polling()
updater.idle()





import telegram
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler

# 填写您的 Telegram 机器人 Token
TOKEN = '5983348762:AAGyhzIL29C0lal3zjgAsH9JeWutoor_W-A'

# 定义处理 /start 命令的函数
def start(update, context):
    # 发送欢迎消息，并附带两个按钮
    message = 'Hello! 点击下面的按钮查看目录'
    keyboard = [[InlineKeyboardButton("按钮1", callback_data='button1'),InlineKeyboardButton("按钮2", url='https://www.example.com')]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    context.bot.send_message(chat_id=update.effective_chat.id, text=message, reply_markup=reply_markup)
    # 设置当前状态为 "button0"
    context.user_data['state'] = 'button0'

# 定义处理按钮点击事件的函数
def button(update, context):
    query = update.callback_query
    data = query.data
    chat_id = query.message.chat_id
    message_id = query.message.message_id
    state = context.user_data.get('state', None)

    if data == 'button1':
        # 点击按钮1，更新消息为“这是按钮1的目录”，并附带两个按钮
        message = '这是按钮1的目录，点击下面的按钮查看内容：'
        keyboard = [
            [InlineKeyboardButton("按钮A", callback_data='buttonA'),InlineKeyboardButton("按钮B", callback_data='buttonB')],
            [InlineKeyboardButton("返回", callback_data='back')]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=message, reply_markup=reply_markup)
        # 设置当前状态为 "button1"
        context.user_data['state'] = 'button1'

    elif data == 'buttonA':
        # 点击按钮A，更新消息为“#这是按钮A的内容”，并附带一个按钮
        message = '#这是按钮A的内容'
        keyboard = [[InlineKeyboardButton("返回", callback_data='back')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        context.bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=message, reply_markup=reply_markup)
        # 设置当前状态为 "buttonA"
        context.user_data['state'] = 'buttonA'

    elif data == 'back':
        # 点击返回按钮，返回到按钮1目录
        if state == 'buttonA':
            # 如果当前在按钮A目录，返回到按钮1目录
            message = '这是按钮1的目录，点击下面的按钮查看内容：'
            keyboard = [
                [InlineKeyboardButton("按钮A", callback_data='buttonA')],
                [InlineKeyboardButton("返回", callback_data='back')]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            context.bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=message, reply_markup=reply_markup)
        # 设置当前状态为 "button1"
        context.user_data['state'] = 'button1'
        
        # 点击返回按钮，返回到上一级目录
        if state == 'button1':
            # 如果当前在按钮1目录，返回到首页
            message = 'Hello! 点击下面的按钮查看目录：'
            keyboard = [[InlineKeyboardButton("按钮1", callback_data='button1'),InlineKeyboardButton("按钮2", url='https://www.example.com')]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            context.bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=message, reply_markup=reply_markup)
            # 设置当前状态为 "button0"
            context.user_data['state'] = 'button0'
            
# 创建一个 Updater 对象，并添加处理函数
updater = Updater(TOKEN, use_context=True)
dispatcher = updater.dispatcher
dispatcher.add_handler(CommandHandler('start', start))
dispatcher.add_handler(CallbackQueryHandler(button))

# 启动机器人
updater.start_polling()